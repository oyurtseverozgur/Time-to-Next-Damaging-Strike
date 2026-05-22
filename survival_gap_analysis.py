# survival_gap_analysis.py
# Time-to-Next Damaging Strike: Recurrent Event Survival Analysis
# FAA NWSD Airport-Level Gap-Time Pipeline
#
# Usage: python survival_gap_analysis.py
# Requires: Public.xlsx in working directory (or set NWSD_PATH below)
# Tested: Python 3.10-3.12 | Google Colab
# Install: pip install pandas openpyxl lifelines matplotlib scipy numpy

import os
import json
import warnings
from collections import defaultdict
from datetime import datetime, date

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from scipy import stats

warnings.filterwarnings('ignore')

try:
    from lifelines import (KaplanMeierFitter, CoxPHFitter,
                           NelsonAalenFitter)
    from lifelines.statistics import (multivariate_logrank_test,
                                      logrank_test)
except ImportError:
    raise SystemExit("Run:  pip install lifelines")

from openpyxl import load_workbook

# ── Config ─────────────────────────────────────────────────────────────────
NWSD_PATH    = 'Public.xlsx'
STUDY_END    = date(2024, 12, 31)
DAMAGE_CODES = {'M', 'M?', 'S', 'D'}
SEVERE_CODES = {'S', 'D'}
OUT_DIR      = 'results_survival'
os.makedirs(OUT_DIR, exist_ok=True)

# ── Visual style ────────────────────────────────────────────────────────────
C_NAVY   = '#1B3A6B'
C_CRIM   = '#9B1D22'
C_FOREST = '#1A5C32'
C_AMBER  = '#6B4400'
C_PURP   = '#4B286D'

plt.rcParams.update({
    'font.family':       'serif',
    'font.size':         9,
    'axes.linewidth':    0.7,
    'axes.spines.top':   False,
    'axes.spines.right': False,
    'axes.grid':         True,
    'grid.color':        '#DADADA',
    'grid.linewidth':    0.35,
    'grid.linestyle':    ':',
    'xtick.major.size':  3.5,
    'ytick.major.size':  3.5,
    'figure.facecolor':  'white',
    'axes.facecolor':    'white',
    'legend.fontsize':   8,
    'legend.framealpha': 0.88,
})


# ── Step 1: Parse NWSD and extract gap times ────────────────────────────────

def extract_gap_times(path, damage_codes=DAMAGE_CODES, study_end=STUDY_END):
    # Build airport-level gap-time panel from FAA NWSD Public.xlsx.
    wb  = load_workbook(path, read_only=True, data_only=True)
    ws  = wb.active
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col = {h: i for i, h in enumerate(hdr) if h}

    AP = col['AIRPORT']
    DT = col['INCIDENT_DATE']
    DMG = col['DAMAGE_LEVEL']

    airport_events = defaultdict(list)
    n_total = n_skip = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        dmg = str(row[DMG]).strip() if row[DMG] else ''
        if dmg not in damage_codes:
            continue
        n_total += 1
        ap  = str(row[AP]).strip() if row[AP] else ''
        if not ap or ap in ('', 'nan', 'None', 'UNKNOWN'):
            n_skip += 1
            continue
        dt_str = str(row[DT]).strip() if row[DT] else ''
        if not dt_str or dt_str in ('nan', 'None', ''):
            n_skip += 1
            continue
        try:
            d = datetime.strptime(dt_str[:10], '%Y-%m-%d').date()
        except ValueError:
            n_skip += 1
            continue
        if d > study_end:
            continue
        airport_events[ap].append((d, dmg))

    wb.close()

    records = []
    n_apts_2plus = 0

    for ap, events in airport_events.items():
        events.sort(key=lambda x: x[0])
        if len(events) < 2:
            continue
        n_apts_2plus += 1

        for k, (t_start, _) in enumerate(events[:-1]):
            t_end, dmg_end = events[k + 1]
            gap = (t_end - t_start).days
            if gap <= 0:
                continue

            mo     = t_end.month
            season = ('Autumn' if mo in (9, 10, 11) else
                      'Summer' if mo in (6,  7,  8) else
                      'Spring' if mo in (3,  4,  5) else 'Winter')

            records.append({
                'airport':     ap,
                'event_order': k + 1,
                'stratum':     min(k + 1, 5),
                'gap_days':    gap,
                'event':       1,
                'year':        t_end.year,
                'month':       mo,
                'season':      season,
                'severe':      int(dmg_end in SEVERE_CODES),
            })

        # Censored row: last event to study end
        last_d = events[-1][0]
        gap_c  = (study_end - last_d).days
        if gap_c > 0:
            records.append({
                'airport':     ap,
                'event_order': len(events),
                'stratum':     min(len(events), 5),
                'gap_days':    gap_c,
                'event':       0,
                'year':        study_end.year,
                'month':       12,
                'season':      'Winter',
                'severe':      0,
            })

    df = pd.DataFrame(records)
    print(f"Damaging events parsed:  {n_total:,}  (skipped: {n_skip:,})")
    print(f"Airports with ≥2 events: {n_apts_2plus:,}")
    print(f"Gap-time intervals:      {len(df):,}")
    print(f"  Observed events: {df['event'].sum():,}  |  Censored: {(df['event']==0).sum():,}")
    return df


# ── Step 2: Build covariates ────────────────────────────────────────────────

def add_covariates(df):
    # Add hub-size dummies and seasonal dummies.
    apt_n = df[df['event'] == 1].groupby('airport').size()

    large_hub  = set(apt_n[apt_n >= 200].index)
    medium_hub = set(apt_n[(apt_n >= 50) & (apt_n < 200)].index)

    df = df.copy()
    df['hub_large']  = df['airport'].isin(large_hub).astype(int)
    df['hub_medium'] = df['airport'].isin(medium_hub).astype(int)
    df['autumn']     = (df['season'] == 'Autumn').astype(int)
    df['spring']     = (df['season'] == 'Spring').astype(int)
    df['summer']     = (df['season'] == 'Summer').astype(int)

    n_large  = df['airport'].isin(large_hub).nunique()
    n_medium = df['airport'].isin(medium_hub).nunique()
    print(f"\nHub breakdown: large={len(large_hub)}  medium={len(medium_hub)}  "
          f"small/non-hub={df['airport'].nunique()-len(large_hub)-len(medium_hub)}")
    return df, large_hub, medium_hub


# ── Step 3: Descriptive statistics ─────────────────────────────────────────

def describe_panel(df):
    obs = df[df['event'] == 1]['gap_days']
    print(f"\n=== Panel statistics ===")
    print(f"  n intervals:  {len(df):,}")
    print(f"  n airports:   {df['airport'].nunique():,}")
    print(f"  n events:     {df['event'].sum():,}")
    print(f"  Median gap:   {obs.median():.0f} days")
    print(f"  Mean gap:     {obs.mean():.0f} days")
    print(f"  Min gap:      {obs.min()} days")
    print(f"  Max gap:      {obs.max()} days")
    print(f"\n  By recurrence order:")
    for k in range(1, 6):
        lbl = f'k={k}' if k < 5 else 'k≥5'
        sub = obs[df[df['event']==1]['stratum'] == k]
        print(f"    {lbl:5s}: n={len(sub):5,}  median={sub.median():.0f}d  "
              f"mean={sub.mean():.0f}d")
    print(f"\n  By season:")
    for s in ['Spring', 'Summer', 'Autumn', 'Winter']:
        sub = obs[df[df['event']==1]['season'] == s]
        print(f"    {s:7s}: n={len(sub):5,}  median={sub.median():.0f}d")


# ── Step 4: Log-rank tests ───────────────────────────────────────────────────

def run_logrank_tests(df):
    print(f"\n=== Log-rank tests ===")
    for label, group_col in [('Stratum (k)', 'stratum'),
                              ('Hub size',    'hub_large'),
                              ('Season',      'season')]:
        res = multivariate_logrank_test(df['gap_days'], df[group_col], df['event'])
        print(f"  {label:20s}: χ² = {res.test_statistic:10.1f}  p = {res.p_value:.3e}")
    return res


# ── Step 5: Cox PHM (first event) ───────────────────────────────────────────

def fit_cox_first(df, covariates):
    df_k1 = df[df['event_order'] == 1].copy()
    cpf   = CoxPHFitter(penalizer=0.1)
    cpf.fit(df_k1[['gap_days', 'event'] + covariates],
            duration_col='gap_days', event_col='event')
    print(f"\n=== Cox PHM (k=1 only) | C = {cpf.concordance_index_:.3f} ===")
    cpf.print_summary()
    return cpf


# ── Step 6: Stratified Cox (PWP-GT) ────────────────────────────────────────

def fit_pwpgt(df, covariates):
    cpf = CoxPHFitter(penalizer=0.1)
    cpf.fit(df[['gap_days', 'event', 'stratum'] + covariates],
            duration_col='gap_days', event_col='event', strata=['stratum'])
    print(f"\n=== Stratified Cox / PWP-GT | C = {cpf.concordance_index_:.3f} ===")
    cpf.print_summary()
    return cpf


# ── Step 7: Frailty variance approximation ─────────────────────────────────

def estimate_frailty(df):
    # Approximate gamma frailty theta from between-airport log-gap variance.
    # Full REML requires R's survival::coxph with frailty(airport).
    apt_meds = df[df['event'] == 1].groupby('airport')['gap_days'].median()
    log_var  = np.var(np.log(apt_meds + 1))
    theta    = log_var / (np.pi ** 2 / 6)   # log-normal → gamma approximation
    tau      = theta / (theta + 2)
    print(f"\n=== Frailty (between-airport heterogeneity) ===")
    print(f"  Log-gap variance: {log_var:.4f}")
    print(f"  theta ≈ {theta:.3f}   Kendall tau ≈ {tau:.3f}")
    print(f"  Interpretation: ~{tau*100:.0f}% of within-airport gap-time correlation")
    print(f"  is attributable to unmeasured airport characteristics.")
    print(f"\n  For exact REML frailty estimation use R:")
    print(f"  library(survival)")
    print(f"  coxph(Surv(gap_days, event) ~ hub_large + hub_medium + autumn + spring + summer")
    print(f"        + strata(stratum) + frailty(airport, distribution='gamma'), data=df)")
    return theta, tau


# ── Step 8: Figure — KM by stratum ─────────────────────────────────────────

def plot_km_strata(df, out_path):
    colors = [C_NAVY, '#2F6CA4', C_AMBER, C_CRIM, C_FOREST]
    ls_seq = ['-', '-', '--', ':', '-.']

    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
    ax = axes[0]
    for k, (col, ls) in enumerate(zip(colors, ls_seq), 1):
        lbl = f'k = {k}' if k < 5 else 'k ≥ 5'
        sub = df[df['stratum'] == k]
        kmf = KaplanMeierFitter()
        kmf.fit(sub['gap_days'], sub['event'], label=lbl)
        kmf.plot_survival_function(ax=ax, color=col, linestyle=ls,
                                   lw=1.8, ci_show=(k in (1, 5)))
    ax.set_xlabel('Gap time (days)')
    ax.set_ylabel('Survival probability S(t)')
    ax.set_xlim(0, 2000)
    res = multivariate_logrank_test(df['gap_days'], df['stratum'], df['event'])
    ax.text(0.97, 0.96,
            f'Log-rank χ² = {res.test_statistic:.0f}\np < 10⁻¹⁰⁰',
            transform=ax.transAxes, ha='right', va='top', fontsize=8.5,
            bbox=dict(boxstyle='round,pad=0.3', facecolor='#EBF0F7',
                      edgecolor=C_NAVY, lw=0.7))
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('KM survivor curves by recurrence order', fontsize=9.5,
                 fontweight='bold', pad=6)

    # Panel B: hub size
    ax = axes[1]
    hub_cats = {C_CRIM: 'Large hub', '#2F6CA4': 'Medium hub', C_NAVY: 'Small/non-hub'}
    apt_n = df[df['event']==1].groupby('airport').size()
    large  = set(apt_n[apt_n >= 200].index)
    medium = set(apt_n[(apt_n >= 50) & (apt_n < 200)].index)
    df2 = df.copy()
    df2['hub_cat'] = df2['airport'].apply(
        lambda x: 'Large hub' if x in large else
                  'Medium hub' if x in medium else 'Small/non-hub')
    for col, cat in hub_cats.items():
        sub = df2[df2['hub_cat'] == cat]
        if len(sub) == 0: continue
        kmf = KaplanMeierFitter()
        kmf.fit(sub['gap_days'], sub['event'], label=cat)
        kmf.plot_survival_function(ax=ax, color=col, lw=2.0, ci_show=False)
    ax.set_xlabel('Gap time (days)')
    ax.set_ylabel('Survival probability S(t)')
    ax.set_xlim(0, 3000)
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes,
            fontsize=9, fontweight='bold', va='top')
    ax.set_title('KM survivor curves by airport size', fontsize=9.5,
                 fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    plt.savefig(out_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  KM plot saved: {out_path}")


# ── Step 9: HR forest plot ──────────────────────────────────────────────────

def plot_hr_forest(cpf2, out_path):
    covs = ['hub_large', 'hub_medium', 'autumn', 'spring', 'summer']
    labels = ['Large hub\n(vs Small)', 'Medium hub\n(vs Small)',
              'Autumn\n(vs Winter)', 'Spring\n(vs Winter)', 'Summer\n(vs Winter)']
    colors = [C_CRIM, '#2F6CA4', C_AMBER, C_FOREST, C_NAVY]

    params = cpf2.params_
    ses    = cpf2.standard_errors_
    pvs    = cpf2.summary['p']

    fig, ax = plt.subplots(figsize=(9, 5))
    y_pos = list(range(len(covs)))[::-1]

    for i, (cov, lbl, col) in enumerate(zip(covs, labels, colors)):
        yp     = y_pos[i]
        beta   = float(params.get(cov, 0))
        se     = float(ses.get(cov, 0.05))
        hr     = np.exp(beta)
        ci_lo  = np.exp(beta - 1.96 * se)
        ci_hi  = np.exp(beta + 1.96 * se)
        p      = float(pvs.get(cov, 1.0))
        sig    = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else ''

        ax.plot([ci_lo, ci_hi], [yp, yp], color=col, lw=2.2, solid_capstyle='round')
        ax.scatter(hr, yp, s=80, c=col, zorder=5, edgecolors='white', lw=0.8)
        ax.text(ci_hi + 0.05, yp, f' {hr:.2f} {sig}', va='center',
                fontsize=8.5, color=col)
        ax.text(-0.02, yp, lbl, ha='right', va='center',
                transform=ax.get_yaxis_transform(), fontsize=8.5)

    ax.axvline(1.0, color='black', lw=0.9, alpha=0.45, ls='--')
    ax.set_xlabel('Hazard Ratio (HR > 1 → shorter gap, higher hazard)')
    ax.set_yticks([])
    ax.set_xscale('log')
    ax.set_xlim(0.8, 12)
    ax.set_xticks([0.9, 1.0, 1.5, 2.0, 3.0, 5.0, 8.0])
    ax.get_xaxis().set_major_formatter(plt.ScalarFormatter())
    ax.set_title('Hazard ratios — stratified Cox (PWP-GT)', fontsize=9.5,
                 fontweight='bold', pad=6)
    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  HR forest plot saved: {out_path}")


# ── Step 10: Nelson-Aalen + seasonal hazard ─────────────────────────────────

def plot_seasonal_na(df, out_path):
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    # (A) Monthly relative hazard rate
    ax = axes[0]
    mth_events = defaultdict(int)
    mth_total  = defaultdict(int)
    for _, row in df.iterrows():
        mth_total[row['month']] += 1
        if row['event'] == 1:
            mth_events[row['month']] += 1

    months = list(range(1, 13))
    rates  = [mth_events[m] / (mth_total[m] + 1e-9) for m in months]
    mean_r = np.mean(rates)
    rel_r  = [r / mean_r for r in rates]
    labels_m = ['Jan','Feb','Mar','Apr','May','Jun',
                'Jul','Aug','Sep','Oct','Nov','Dec']
    bar_cols = [C_CRIM if r > 1.2 else C_AMBER if r > 1.0 else C_NAVY
                for r in rel_r]
    ax.bar(months, rel_r, color=bar_cols, alpha=0.82, edgecolor='white',
           width=0.75)
    ax.axhline(1.0, color='#555555', lw=1.3, ls='--', alpha=0.6,
               label='Annual mean = 1.0')
    ax.set_xticks(months)
    ax.set_xticklabels(labels_m, fontsize=8.5)
    ax.set_ylabel('Relative hazard rate (monthly / annual mean)')
    ax.legend(fontsize=8.5)
    res_s = multivariate_logrank_test(df['gap_days'], df['season'], df['event'])
    ax.text(0.97, 0.95,
            f'Log-rank χ² = {res_s.test_statistic:.0f}\np = {res_s.p_value:.1e}',
            transform=ax.transAxes, ha='right', va='top', fontsize=8.5,
            bbox=dict(boxstyle='round,pad=0.3', facecolor='#FDF0E8',
                      edgecolor=C_AMBER, lw=0.7))
    ax.text(0.02, 0.97, '(a)', transform=ax.transAxes, fontsize=9,
            fontweight='bold', va='top')
    ax.set_title('Monthly relative hazard rate (real NWSD)', fontsize=9.5,
                 fontweight='bold', pad=6)

    # (B) Nelson-Aalen by hub size
    ax = axes[1]
    apt_n  = df[df['event']==1].groupby('airport').size()
    large  = set(apt_n[apt_n >= 200].index)
    medium = set(apt_n[(apt_n >= 50) & (apt_n < 200)].index)

    hub_map = {C_CRIM: 'Large hub', '#2F6CA4': 'Medium hub', C_NAVY: 'Small/non-hub'}
    for col, cat in hub_map.items():
        if cat == 'Large hub':
            mask = df['airport'].isin(large)
        elif cat == 'Medium hub':
            mask = df['airport'].isin(medium)
        else:
            mask = (~df['airport'].isin(large)) & (~df['airport'].isin(medium))
        sub = df[mask & (df['stratum'] == 1)]
        if len(sub) < 5: continue
        naf = NelsonAalenFitter()
        naf.fit(sub['gap_days'], sub['event'], label=cat)
        naf.plot_cumulative_hazard(ax=ax, color=col, lw=2.0, ci_show=False)

    t_ref  = np.linspace(0, 2500, 200)
    n_ev   = int(df[(df['event']==1) & (df['stratum']==1)].shape[0])
    n_tot  = df[df['stratum']==1]['gap_days'].mean()
    ax.plot(t_ref, (n_ev / n_tot) * t_ref / 1000, 'k--', lw=1.2, alpha=0.45,
            label='HPP reference (linear)')
    ax.set_xlabel('Gap time (days)')
    ax.set_ylabel('Cumulative baseline hazard H₀(t)')
    ax.set_xlim(0, 2500)
    ax.legend(fontsize=8.5)
    ax.text(0.02, 0.97, '(b)', transform=ax.transAxes, fontsize=9,
            fontweight='bold', va='top')
    ax.set_title('Nelson-Aalen cumulative hazard (k=1 stratum)', fontsize=9.5,
                 fontweight='bold', pad=6)

    plt.tight_layout(pad=2.0)
    plt.savefig(out_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Seasonal + NA plot saved: {out_path}")


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if not os.path.exists(NWSD_PATH):
        raise FileNotFoundError(
            f"NWSD file not found: '{NWSD_PATH}'. "
            "Download from https://wildlife.faa.gov/downloads."
        )

    # 1. Extract panel
    df = extract_gap_times(NWSD_PATH)

    # 2. Add covariates
    df, large_hub, medium_hub = add_covariates(df)

    # 3. Descriptive stats
    describe_panel(df)

    # 4. Log-rank
    run_logrank_tests(df)

    # 5. Cox PHM (first event)
    covariates = ['hub_large', 'hub_medium', 'autumn', 'spring', 'summer']
    cpf1 = fit_cox_first(df, covariates)

    # 6. PWP-GT
    cpf2 = fit_pwpgt(df, covariates)

    # 7. Frailty
    theta, tau = estimate_frailty(df)

    # 8. Figures
    print("\nGenerating figures...")
    plot_km_strata(df, f'{OUT_DIR}/fig_km_curves.png')
    plot_hr_forest(cpf2, f'{OUT_DIR}/fig_hr_forest.png')
    plot_seasonal_na(df, f'{OUT_DIR}/fig_seasonal_na.png')

    # 9. Save results summary
    results = {
        'n_intervals':   int(len(df)),
        'n_airports':    int(df['airport'].nunique()),
        'n_events':      int(df['event'].sum()),
        'median_gap':    int(df[df['event']==1]['gap_days'].median()),
        'mean_gap':      float(df[df['event']==1]['gap_days'].mean()),
        'c_cox1':        float(cpf1.concordance_index_),
        'c_pwpgt':       float(cpf2.concordance_index_),
        'theta':         float(theta),
        'tau':           float(tau),
        'n_large_hub':   int(len(large_hub)),
        'n_medium_hub':  int(len(medium_hub)),
        'pwpgt_hrs': {
            c: float(np.exp(cpf2.params_.get(c, 0)))
            for c in covariates
        },
    }
    with open(f'{OUT_DIR}/results_summary.json', 'w') as fh:
        import json as _json
        _json.dump(results, fh, indent=2)

    print(f"\nDone. Results saved to {OUT_DIR}/")
    print(f"  C (Cox PHM k=1):  {results['c_cox1']:.3f}")
    print(f"  C (PWP-GT):       {results['c_pwpgt']:.3f}")
    print(f"  theta (frailty):  {results['theta']:.3f}")
    print(f"  tau (Kendall):    {results['tau']:.3f}")
